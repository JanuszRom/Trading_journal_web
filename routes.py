from flask import request, jsonify, current_app, send_from_directory
from werkzeug.utils import secure_filename
from models import db, Trade, Screenshot
import os
import uuid
from datetime import datetime


def register_routes(app):
    # Get all trades
    @app.route('/api/trades', methods=['GET'])
    def get_trades():
        trades = Trade.query.order_by(Trade.timestamp.desc()).all()

        result = []
        for trade in trades:
            trade_data = {
                'id': trade.id,
                'timestamp': trade.timestamp.isoformat(),
                'instrument': trade.instrument,
                'direction': trade.direction,
                'entry': trade.entry,
                'exit': trade.exit,
                'stop_loss': trade.stop_loss,
                'take_profit': trade.take_profit,
                'size': trade.size,
                'risk': trade.risk,
                'reward': trade.reward,
                'profit_loss': trade.profit_loss,
                'duration': trade.duration,
                'strategy': trade.strategy,
                'setup': trade.setup,
                'mistakes': trade.mistakes,
                'lessons': trade.lessons,
                'screenshots': [{'id': s.id, 'filename': s.filename} for s in trade.screenshots]
            }
            result.append(trade_data)

        return jsonify(result)

    # Get a specific trade
    @app.route('/api/trades/<int:trade_id>', methods=['GET'])
    def get_trade(trade_id):
        trade = Trade.query.get_or_404(trade_id)

        trade_data = {
            'id': trade.id,
            'timestamp': trade.timestamp.isoformat(),
            'instrument': trade.instrument,
            'direction': trade.direction,
            'entry': trade.entry,
            'exit': trade.exit,
            'stop_loss': trade.stop_loss,
            'take_profit': trade.take_profit,
            'size': trade.size,
            'risk': trade.risk,
            'reward': trade.reward,
            'profit_loss': trade.profit_loss,
            'duration': trade.duration,
            'strategy': trade.strategy,
            'setup': trade.setup,
            'mistakes': trade.mistakes,
            'lessons': trade.lessons,
            'screenshots': [{'id': s.id, 'filename': s.filename} for s in trade.screenshots]
        }

        return jsonify(trade_data)

    # Create a new trade
    @app.route('/api/trades', methods=['POST'])
    def create_trade():
        data = request.form

        # Create new trade
        new_trade = Trade(
            instrument=data.get('instrument'),
            direction=data.get('direction'),
            entry=float(data.get('entry')),
            exit=float(data.get('exit')),
            stop_loss=float(data.get('stop_loss')),
            take_profit=float(data.get('take_profit')),
            size=float(data.get('size')),
            risk=float(data.get('risk')),
            reward=float(data.get('reward')),
            profit_loss=float(data.get('profit_loss')),
            duration=data.get('duration'),
            strategy=data.get('strategy'),
            setup=data.get('setup', ''),
            mistakes=data.get('mistakes', ''),
            lessons=data.get('lessons', '')
        )

        db.session.add(new_trade)
        db.session.commit()

        # Handle screenshot uploads
        files = request.files.getlist('screenshots')
        for file in files:
            if file and file.filename:
                # Secure filename and make unique
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4().hex}_{filename}"

                # Save file
                filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
                file.save(filepath)

                # Create screenshot record
                screenshot = Screenshot(
                    filename=unique_filename,
                    filepath=filepath,
                    trade_id=new_trade.id
                )
                db.session.add(screenshot)

        db.session.commit()

        return jsonify({'success': True, 'trade_id': new_trade.id}), 201

    # Update a trade
    @app.route('/api/trades/<int:trade_id>', methods=['PUT'])
    def update_trade(trade_id):
        trade = Trade.query.get_or_404(trade_id)
        data = request.form

        # Update trade fields
        trade.instrument = data.get('instrument', trade.instrument)
        trade.direction = data.get('direction', trade.direction)
        trade.entry = float(data.get('entry', trade.entry))
        trade.exit = float(data.get('exit', trade.exit))
        trade.stop_loss = float(data.get('stop_loss', trade.stop_loss))
        trade.take_profit = float(data.get('take_profit', trade.take_profit))
        trade.size = float(data.get('size', trade.size))
        trade.risk = float(data.get('risk', trade.risk))
        trade.reward = float(data.get('reward', trade.reward))
        trade.profit_loss = float(data.get('profit_loss', trade.profit_loss))
        trade.duration = data.get('duration', trade.duration)
        trade.strategy = data.get('strategy', trade.strategy)
        trade.setup = data.get('setup', trade.setup)
        trade.mistakes = data.get('mistakes', trade.mistakes)
        trade.lessons = data.get('lessons', trade.lessons)

        # Handle new screenshots
        files = request.files.getlist('screenshots')
        for file in files:
            if file and file.filename:
                # Secure filename and make unique
                filename = secure_filename(file.filename)
                unique_filename = f"{uuid.uuid4().hex}_{filename}"

                # Save file
                filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
                file.save(filepath)

                # Create screenshot record
                screenshot = Screenshot(
                    filename=unique_filename,
                    filepath=filepath,
                    trade_id=trade.id
                )
                db.session.add(screenshot)

        db.session.commit()

        return jsonify({'success': True})

    # Delete a trade
    @app.route('/api/trades/<int:trade_id>', methods=['DELETE'])
    def delete_trade(trade_id):
        trade = Trade.query.get_or_404(trade_id)

        # Delete associated screenshots
        for screenshot in trade.screenshots:
            # Delete file from disk
            try:
                os.remove(screenshot.filepath)
            except:
                pass

            # Delete record
            db.session.delete(screenshot)

        # Delete trade
        db.session.delete(trade)
        db.session.commit()

        return jsonify({'success': True})

    # Get screenshot
    @app.route('/api/screenshots/<filename>')
    def get_screenshot(filename):
        return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)

    # Delete screenshot
    @app.route('/api/screenshots/<int:screenshot_id>', methods=['DELETE'])
    def delete_screenshot(screenshot_id):
        screenshot = Screenshot.query.get_or_404(screenshot_id)

        # Delete file from disk
        try:
            os.remove(screenshot.filepath)
        except:
            pass

        # Delete record
        db.session.delete(screenshot)
        db.session.commit()

        return jsonify({'success': True})

    # Export to Excel endpoint
    @app.route('/api/export/excel')
    def export_excel():
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from flask import send_file

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Trade Journal"

        # Add headers
        headers = [
            'Timestamp', 'Instrument', 'Direction', 'Entry', 'Exit',
            'Stop Loss', 'Take Profit', 'Size', 'Risk', 'Reward',
            'P/L', 'Duration', 'Strategy', 'Setup', 'Mistakes',
            'Lessons'
        ]
        ws.append(headers)

        # Format headers
        header_style = {
            'fill': PatternFill(start_color='4F81BD', fill_type='solid'),
            'font': Font(color='FFFFFF', bold=True),
            'alignment': Alignment(horizontal='center')
        }

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            for attr, value in header_style.items():
                setattr(cell, attr, value)

        # Add data
        trades = Trade.query.all()
        for idx, trade in enumerate(trades):
            row = [
                trade.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                trade.instrument,
                trade.direction,
                trade.entry,
                trade.exit,
                trade.stop_loss,
                trade.take_profit,
                trade.size,
                trade.risk,
                trade.reward,
                trade.profit_loss,
                trade.duration,
                trade.strategy,
                trade.setup,
                trade.mistakes,
                trade.lessons
            ]
            ws.append(row)

        # Adjust column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Return file for download
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name='trades_export.xlsx',
            as_attachment=True
        )